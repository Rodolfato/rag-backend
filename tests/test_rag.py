import os
import time
from typing import Any, Dict, List
from openpyxl import load_workbook
from app.engines.engine_interface import Engine
from app.services.llm_services import query_llm
from app.config import vector_db_engine, llm
from langchain_core.language_models import BaseLLM

from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
from nltk.translate.meteor_score import meteor_score
from nltk.tokenize import word_tokenize

SHEET_FILE_PATH = "tests/sheets/evaluacion.xlsx"
OUTPUT_SHEET_FILE_PATH = "tests/sheets/"
SHEET_STRUCTURE = [
    "question",
    "expected_answer",
    "actual_answers",
    "response_times",
    "cosine_similarity",
    "meteor_score",
    "subjective_evaluation",
]


def get_questions_and_expected_answers_from_excel(
    filename: str, sheet: str, sheet_structure: list
) -> List[Dict]:
    workbook = load_workbook(filename=filename)
    sheet = workbook[sheet]
    questions = []

    for row in sheet.iter_rows(min_row=2, max_col=2):
        sheet_dict = {field: None for field in sheet_structure}

        first_cell_value = row[0].value
        second_cell_value = row[1].value

        if first_cell_value is not None and second_cell_value is not None:
            sheet_dict[sheet_structure[0]] = first_cell_value
            sheet_dict[sheet_structure[1]] = second_cell_value
            questions.append(sheet_dict)

    return questions


def get_sheet_names(filename: str):
    workbook = load_workbook(filename=filename)
    sheet_names = workbook.sheetnames
    return sheet_names[:-1]


def get_answers(
    vector_model,
    question_dict: Dict[str, Any],
    repetitions: int = 5,
    vector_db_engine: Engine = None,
    llm: BaseLLM = None,
    top_k: int = 10,
) -> Dict[str, Any]:
    answers = []
    response_times = []
    cosine_similarity_scores = []
    meteor_scores = []
    for _ in range(repetitions):
        try:
            start_time = time.time()
            answer = query_llm(vector_db_engine, question_dict["question"], llm, top_k)
            end_time = time.time()

            cos_sim = calculate_cosine_similarity(
                question_dict["expected_answer"],
                answer.get("model_response"),
                vector_model,
            )
            met_score = calculate_meteor_score(
                question_dict["expected_answer"], answer.get("model_response")
            )

            answers.append(answer.get("model_response", None).strip())
            response_times.append(end_time - start_time)
            cosine_similarity_scores.append(cos_sim)
            meteor_scores.append(met_score)

        except Exception as e:
            print(f"Error querying LLM: {e}")
            answers.append(None)

    result_dict = question_dict.copy()
    result_dict["actual_answers"] = answers
    result_dict["response_times"] = response_times
    result_dict["cosine_similarity"] = cosine_similarity_scores
    result_dict["meteor_score"] = meteor_scores

    return result_dict


def write_answers_to_xlsx(
    question_dict: Dict, filename: str, sheet: str, output_path: str, row_start
) -> None:
    workbook = load_workbook(filename=filename)
    sheet = workbook[sheet]
    # Se escriben las respuestas obtenidas
    column_index = 3
    for row_index, value in enumerate(
        question_dict["actual_answers"], start=2 + row_start
    ):
        print(
            f"Escribiendo respuesta del proyecto {sheet} en ({row_index},{column_index}) "
        )
        sheet.cell(row=row_index, column=column_index, value=value)
    # Se escriben los tiempos de respuesta
    column_index = 4
    for row_index, value in enumerate(
        question_dict["response_times"], start=2 + row_start
    ):
        print(
            f"Escribiendo tiempo de respuesta: {value} del proyecto {sheet} en ({row_index},{column_index}) "
        )
        sheet.cell(row=row_index, column=column_index, value=value)
    # Se escriben las similitudes del coseno
    column_index = 5
    for row_index, value in enumerate(
        question_dict["cosine_similarity"], start=2 + row_start
    ):
        print(
            f"Escribiendo similitud del coseno: {value} del proyecto {sheet} en ({row_index},{column_index}) "
        )
        sheet.cell(row=row_index, column=column_index, value=value)
    # Se escriben los meteor scores
    column_index = 6
    for row_index, value in enumerate(
        question_dict["meteor_score"], start=2 + row_start
    ):
        print(
            f"Escribiendo meteor score: {value} del proyecto {sheet} en ({row_index},{column_index}) "
        )
        sheet.cell(row=row_index, column=column_index, value=value)
    # full_output_path = os.path.join(output_path, "output.xlsx")
    workbook.save(output_path)


def calculate_cosine_similarity(expected_answer, actual_answer, model):
    embeddings = model.encode([actual_answer, expected_answer])

    similarity = cosine_similarity([embeddings[0]], [embeddings[1]])
    return f"{similarity[0][0]:.4f}"


def calculate_meteor_score(expected_answer, actual_answer):
    reference = word_tokenize(expected_answer)
    candidate = word_tokenize(actual_answer)
    score = meteor_score([reference], candidate)
    return f"{score:.4f}"


def print_dict_with_keys_and_values(dictionary: dict):
    for key, value in dictionary.items():
        print(f"{key}: {value}")


model = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")
sheet_names = get_sheet_names(filename=SHEET_FILE_PATH)
start_time = time.time()
for project in sheet_names:
    questions = get_questions_and_expected_answers_from_excel(
        SHEET_FILE_PATH, project, SHEET_STRUCTURE
    )
    index = 0
    for question in questions:
        print(f"El index es ahora {index}")
        answers = get_answers(model, question, 8, vector_db_engine, llm, 10)
        write_answers_to_xlsx(answers, SHEET_FILE_PATH, project, SHEET_FILE_PATH, index)
        index += 8

end_time = time.time()

elapsed_time = end_time - start_time
print(f"Tiempo demorado en hacer el testing: {elapsed_time:.2f} segundos")
